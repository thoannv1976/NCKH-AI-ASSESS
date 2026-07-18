"""Phân hệ báo cáo: dashboard tiến độ, phân loại, xuất báo cáo, hồ sơ năng lực."""
from __future__ import annotations

import io

from fastapi import APIRouter, Depends, HTTPException, Request
from fastapi.responses import Response

from app.auth import require_role
from app.config import ROLE_ADMIN, ROLE_COUNCIL, ROLE_LECTURER, get_settings, now_vn
from app.rubric import all_class_labels, get_rubric, graded_parts, rubric_for, top_class_keys
from app.services.classify import classify

router = APIRouter(prefix="/reports")
staff_dep = Depends(require_role(ROLE_COUNCIL, ROLE_ADMIN))


def render(request: Request, template: str, user: dict, **ctx):
    return request.app.state.templates.TemplateResponse(request, template, {"user": user, "now": now_vn(), **ctx})


def scores_rows(store, khoa: str = "") -> list[dict]:
    """Bảng điểm tất cả hồ sơ đã nộp — điểm HIỆN TẠI (Hội đồng điều chỉnh nếu có, ngược lại điểm AI),
    xem được kể cả khi Hội đồng chưa chốt. Mỗi hồ sơ chỉ truy vấn scores một lần."""
    from collections import defaultdict

    users = {u["id"]: u for u in store.all("users")}
    reviews = {r["submission_id"]: r for r in store.all("reviews")}
    cols = graded_parts(get_rubric(store))  # cột hiển thị theo loại mặc định
    rows = []
    for s in store.all("submissions"):
        if s.get("status") == "draft":
            continue
        u = users.get(s["user_id"])
        if not u:
            continue
        if khoa and (u.get("khoa") or "") != khoa:
            continue
        sub_rubric = rubric_for(store, s)
        part_max = {p: pd["max_score"] for p, pd in sub_rubric.get("parts", {}).items()}
        scores = store.find("scores", submission_id=s["id"])
        has = bool(scores)
        part_sum: dict[str, float] = defaultdict(float)
        adjusted = False
        for sc in scores:
            part_sum[sc["part"]] += sc.get("final_score") or 0
            if sc.get("council_score") is not None:
                adjusted = True
        part_totals = {p: round(min(v, part_max.get(p, v)), 2) for p, v in part_sum.items()}
        totals = {p: part_totals.get(p, "") for p in cols}  # map về cột hiển thị
        total = round(sum(part_totals.values()), 2) if has else None
        review = reviews.get(s["id"]) or {}
        final = review.get("status") in ("approved", "published")
        level = None
        if total is not None:
            # nếu đã chốt thì dùng phân loại đã lưu, ngược lại tính tạm theo điểm hiện tại
            level = review.get("classification_label") if final else classify(total, sub_rubric)["label"]
        rows.append({
            "user": u, "sub": s, "totals": totals, "total": total, "has": has,
            "level": level, "final": final, "adjusted": adjusted,
            "ai_total": s.get("ai_total"),
        })
    rows.sort(key=lambda r: (-(r["total"] if r["total"] is not None else -1),
                             r["user"].get("khoa", ""), r["user"].get("ma_gv", "")))
    return rows


def build_summary(store) -> dict:
    users = {u["id"]: u for u in store.all("users") if u["role"] == ROLE_LECTURER}
    subs = store.all("submissions")
    reviews = {r["submission_id"]: r for r in store.all("reviews")}

    by_khoa: dict[str, dict] = {}
    for u in users.values():
        k = u.get("khoa") or "(Chưa rõ đơn vị)"
        by_khoa.setdefault(k, {"khoa": k, "total": 0, "submitted": 0, "graded": 0, "approved": 0, "published": 0})
        by_khoa[k]["total"] += 1
    class_labels = all_class_labels(store)
    classification = {key: 0 for key in class_labels}
    top_keys = top_class_keys(store)
    from collections import defaultdict
    part_sums: dict[str, list] = defaultdict(list)
    core_list = []

    for s in subs:
        u = users.get(s["user_id"])
        if not u:
            continue
        k = u.get("khoa") or "(Chưa rõ đơn vị)"
        st = s.get("status")
        if st in ("submitted", "locked", "grading", "graded", "approved", "published"):
            by_khoa[k]["submitted"] += 1
        if st in ("graded", "approved", "published"):
            by_khoa[k]["graded"] += 1
        if st in ("approved", "published"):
            by_khoa[k]["approved"] += 1
        if st == "published":
            by_khoa[k]["published"] += 1
        r = reviews.get(s["id"])
        if r and r.get("classification"):
            classification[r["classification"]] = classification.get(r["classification"], 0) + 1
            for p, v in (r.get("part_totals") or {}).items():
                part_sums[p].append(v)
            if r["classification"] in top_keys:
                core_list.append({"user": u, "total": r["total_final"], "submission_id": s["id"]})

    cols = graded_parts(get_rubric(store))
    part_avgs = {p: (round(sum(part_sums[p]) / len(part_sums[p]), 2) if part_sums.get(p) else 0) for p in cols}
    core_list.sort(key=lambda x: -x["total"])
    return {
        "by_khoa": sorted(by_khoa.values(), key=lambda x: x["khoa"]),
        "classification": classification, "class_labels": class_labels,
        "part_avgs": part_avgs, "core_list": core_list,
        "lecturers": len(users),
        "submitted": sum(k["submitted"] for k in by_khoa.values()),
        "graded": sum(k["graded"] for k in by_khoa.values()),
        "published": sum(k["published"] for k in by_khoa.values()),
    }


@router.get("")
def dashboard(request: Request, user: dict = staff_dep):
    store = request.app.state.store
    return render(request, "reports/dashboard.html", user, summary=build_summary(store))


@router.get("/scores")
def scores_page(request: Request, khoa: str = "", user: dict = staff_dep):
    store = request.app.state.store
    rows = scores_rows(store, khoa=khoa)
    users = {u["id"]: u for u in store.all("users")}
    khoas = sorted({u.get("khoa", "") for u in users.values() if u.get("khoa")})
    graded = sum(1 for r in rows if r["has"])
    rubric = get_rubric(store)
    return render(request, "reports/scores.html", user, rows=rows, khoas=khoas, khoa=khoa,
                  rubric=rubric, parts=graded_parts(rubric), graded=graded)


@router.get("/scores.xlsx")
def scores_xlsx(request: Request, khoa: str = "", user: dict = staff_dep):
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    store = request.app.state.store
    rubric = get_rubric(store)
    cols = graded_parts(rubric)
    rows = scores_rows(store, khoa=khoa)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bảng điểm"
    header = ["MSSV chủ nhiệm", "Chủ nhiệm", "Email", "Đơn vị", "Bộ môn/Ngành", "Vai trò", "Trạng thái"]
    header += [f"Phần {p} ({rubric['parts'][p]['max_score']})" for p in cols]
    header += ["Tổng (hiện tại)", "Xếp loại", "Tính chất điểm"]
    ws.append(header)
    for c in ws[1]:
        c.fill = PatternFill("solid", fgColor="EA580C")
        c.font = Font(bold=True, color="FFFFFF")
    status_labels = {"submitted": "Đã nộp", "locked": "Đã khóa", "grading": "Đang chấm",
                     "graded": "Đã chấm", "approved": "Đã duyệt", "published": "Đã công bố"}
    for r in rows:
        u = r["user"]
        nature = "Chính thức (đã duyệt)" if r["final"] else ("Tạm tính (chưa chốt)" if r["has"] else "Chưa chấm")
        ws.append([
            u.get("ma_gv", ""), u.get("ho_ten", ""), u.get("email", ""),
            u.get("khoa", ""), u.get("bo_mon", ""), u.get("chuc_vu", ""),
            status_labels.get(r["sub"].get("status"), r["sub"].get("status")),
            *[(r["totals"].get(p, "") if r["has"] else "") for p in cols],
            (r["total"] if r["total"] is not None else ""), (r["level"] or ""), nature,
        ])
    for col, w in zip("ABCDEF", [10, 24, 28, 22, 22, 18]):
        ws.column_dimensions[col].width = w
    buf = io.BytesIO()
    wb.save(buf)
    return Response(
        buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":
                 f"attachment; filename={get_settings().org_short}-BangDiem-{now_vn():%Y%m%d-%H%M}.xlsx"},
    )


@router.get("/phieu-danh-gia.zip")
def reports_zip(request: Request, khoa: str = "", fmt: str = "pdf", user: dict = staff_dep):
    """Tải hàng loạt phiếu đánh giá (ZIP) cho tất cả GV đã chấm (lọc theo đơn vị)."""
    from urllib.parse import quote

    from fastapi.responses import StreamingResponse

    from app.services.report import stream_reports_zip

    store = request.app.state.store
    users = {u["id"]: u for u in store.all("users")}
    subs = [s for s in store.all("submissions")
            if s.get("status") != "draft" and users.get(s["user_id"])
            and (not khoa or (users[s["user_id"]].get("khoa") or "") == khoa)]
    fmt = "docx" if fmt == "docx" else "pdf"
    tag = f"-{khoa}" if khoa else ""
    fname = f"{get_settings().org_short}-PhieuDanhGia{tag}-{now_vn():%Y%m%d}.zip"
    return StreamingResponse(stream_reports_zip(store, subs, fmt), media_type="application/zip", headers={
        "Content-Disposition": f"attachment; filename*=UTF-8''{quote(fname)}",
        "X-Accel-Buffering": "no",
    })


@router.get("/api/summary")
def api_summary(request: Request, user: dict = staff_dep):
    return build_summary(request.app.state.store)


@router.get("/export.xlsx")
def export_xlsx(request: Request, user: dict = staff_dep):
    import openpyxl

    store = request.app.state.store
    rubric = get_rubric(store)
    cols = graded_parts(rubric)
    users = {u["id"]: u for u in store.all("users")}
    reviews = {r["submission_id"]: r for r in store.all("reviews")}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tổng hợp"
    header = ["MSSV chủ nhiệm", "Chủ nhiệm", "Email", "Đơn vị", "Bộ môn/Ngành", "Vai trò", "Trạng thái"]
    header += [f"Phần {p}" for p in cols] + ["Tổng điểm", "Xếp loại"]
    ws.append(header)
    for s in sorted(store.all("submissions"), key=lambda x: -(x.get("final_total") or x.get("ai_total") or 0)):
        u = users.get(s["user_id"], {})
        r = reviews.get(s["id"]) or {}
        pt = r.get("part_totals") or {}
        ws.append([
            u.get("ma_gv", ""), u.get("ho_ten", ""), u.get("email", ""), u.get("khoa", ""), u.get("bo_mon", ""),
            u.get("chuc_vu", ""), s.get("status", ""), *[pt.get(p, "") for p in cols],
            r.get("total_final", ""), r.get("classification_label", ""),
        ])

    ws2 = wb.create_sheet("Xếp loại")
    ws2.append(["Xếp loại", "Số lượng", "Ghi chú"])
    summary = build_summary(store)
    labels = all_class_labels(store)
    for key, label in labels.items():
        note = ""
        for tdef in rubric.get("classification", []):
            if tdef["key"] == key:
                note = tdef.get("note", "")
        ws2.append([label, summary["classification"].get(key, 0), note])

    buf = io.BytesIO()
    wb.save(buf)
    return Response(
        buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":
                 f"attachment; filename={get_settings().org_short}-BaoCao-{now_vn():%Y%m%d}.xlsx"},
    )


@router.get("/profile/{sid}")
def profile(sid: str, request: Request,
            user: dict = Depends(require_role(ROLE_LECTURER, ROLE_COUNCIL, ROLE_ADMIN))):
    """Hồ sơ đánh giá công trình NCKH — trang in được (Ctrl+P → PDF)."""
    store = request.app.state.store
    sub = store.get("submissions", sid)
    if not sub:
        raise HTTPException(404)
    if user["role"] == ROLE_LECTURER:
        if sub["user_id"] != user["id"] or sub.get("status") != "published":
            raise HTTPException(403)
    owner = store.get("users", sub["user_id"])
    review = store.get("reviews", sid)
    rubric = rubric_for(store, sub)
    scores = store.find("scores", submission_id=sid)
    by_part: dict[str, list] = {}
    for s in sorted(scores, key=lambda x: x["criterion_id"]):
        by_part.setdefault(s["part"], []).append(s)
    level_note = ""
    if review and review.get("classification"):
        for c in rubric["classification"]:
            if c["key"] == review["classification"]:
                level_note = c["note"]
    return render(request, "reports/profile.html", user, sub=sub, owner=owner, review=review,
                  rubric=rubric, scores=by_part, parts=graded_parts(rubric), level_note=level_note)
