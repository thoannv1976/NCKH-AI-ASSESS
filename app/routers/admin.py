"""Phân hệ quản trị: danh sách giảng viên, cấu hình, công bố, tải sản phẩm, sao lưu."""
from __future__ import annotations

import csv
import io
import time

from fastapi import APIRouter, Depends, Form, HTTPException, Request, UploadFile
from fastapi.responses import RedirectResponse

from app.auth import require_role, set_password
from app.config import ROLE_ADMIN, ROLE_COUNCIL, ROLE_LECTURER, get_settings, now_vn, parse_dt
from app.rubric import get_rubric, load_rubric_seed, reload_rubric, rubric_types
from app.security import hash_password
from app.services import audit
from app.services.ops import (
    get_timeline, lock_all, publish_all, reset_stuck_grading, send_reminders, unlock_all,
)

router = APIRouter(prefix="/admin")
admin_dep = Depends(require_role(ROLE_ADMIN))


def render(request: Request, template: str, user: dict, **ctx):
    return request.app.state.templates.TemplateResponse(request, template, {"user": user, "now": now_vn(), **ctx})


@router.get("")
def index(request: Request, user: dict = admin_dep):
    from app.services.ai_config import get_ai_config

    store = request.app.state.store
    subs = store.all("submissions")
    by_status: dict[str, int] = {}
    for s in subs:
        by_status[s.get("status", "?")] = by_status.get(s.get("status", "?"), 0) + 1
    users = store.all("users")
    counts = {
        "lecturers": sum(1 for u in users if u["role"] == ROLE_LECTURER),
        "council": sum(1 for u in users if u["role"] == ROLE_COUNCIL),
        "submissions": len(subs),
        "by_status": by_status,
        "graded": by_status.get("graded", 0) + by_status.get("approved", 0) + by_status.get("published", 0),
        "appeals_open": len(store.find("appeals", status="open")),
    }
    settings = get_settings()
    return render(request, "admin/index.html", user, counts=counts, timeline=get_timeline(store),
                  settings=settings, ai_cfg=get_ai_config(store))


@router.get("/tracking")
def tracking_page(request: Request, khoa: str = "", state: str = "", page: int = 1, user: dict = admin_dep):
    from app.services.ops import submission_tracking

    store = request.app.state.store
    rows, counts = submission_tracking(store)
    khoas = sorted({r["user"].get("khoa", "") for r in rows if r["user"].get("khoa")})
    if khoa:
        rows = [r for r in rows if (r["user"].get("khoa") or "") == khoa]
    if state in ("submitted", "draft", "none"):
        rows = [r for r in rows if r["state"] == state]
    total = len(rows)
    per_page = 50
    pages = max(1, (total + per_page - 1) // per_page)
    page = max(1, min(page, pages))
    start = (page - 1) * per_page
    return render(request, "admin/tracking.html", user, rows=rows[start:start + per_page], counts=counts,
                  khoas=khoas, khoa=khoa, state=state, total=total, page=page, pages=pages, per_page=per_page)


@router.get("/tracking.xlsx")
def tracking_xlsx(request: Request, khoa: str = "", state: str = "", user: dict = admin_dep):
    import openpyxl
    from fastapi.responses import Response
    from openpyxl.styles import Font, PatternFill

    from app.services.ops import submission_tracking

    store = request.app.state.store
    rows, _ = submission_tracking(store)
    if khoa:
        rows = [r for r in rows if (r["user"].get("khoa") or "") == khoa]
    if state in ("submitted", "draft", "none"):
        rows = [r for r in rows if r["state"] == state]
    labels = {"submitted": "Đã nộp", "draft": "Bản nháp", "none": "Chưa tạo hồ sơ"}
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TinhTrangNop"
    ws.append(["Mã GV", "Họ tên", "Email", "Đơn vị", "Trạng thái", "Ghi chú / Còn thiếu", "Cập nhật gần nhất"])
    for c in ws[1]:
        c.fill = PatternFill("solid", fgColor="EA580C")
        c.font = Font(bold=True, color="FFFFFF")
    for r in rows:
        u = r["user"]
        if r["state"] == "draft":
            note = "Đủ bài — CHƯA bấm Nộp" if r["ready"] else ("Thiếu: " + ", ".join(r["missing"]))
        elif r["state"] == "none":
            note = "Chưa tạo hồ sơ"
        else:
            note = ""
        ws.append([u.get("ma_gv", ""), u.get("ho_ten", ""), u.get("email", ""), u.get("khoa", ""),
                   labels[r["state"]], note, (r["last"] or "")[:16].replace("T", " ")])
    for col, w in zip("ABCDEFG", [12, 24, 28, 24, 14, 50, 18]):
        ws.column_dimensions[col].width = w
    buf = io.BytesIO()
    wb.save(buf)
    return Response(
        buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":
                 f"attachment; filename={get_settings().org_short}-TinhTrangNop-{now_vn():%Y%m%d}.xlsx"},
    )


@router.post("/lock")
def lock(request: Request, user: dict = admin_dep):
    result = lock_all(request.app.state.store, user)
    return RedirectResponse(f"/admin?locked={result['locked']}&invalid={result['invalid']}", status_code=303)


@router.post("/unlock")
def unlock(request: Request, user: dict = admin_dep):
    """Mở khóa hồ sơ đã khóa để giảng viên nộp/sửa lại (dùng khi gia hạn nộp bài)."""
    n = unlock_all(request.app.state.store, user)
    return RedirectResponse(f"/admin?unlocked={n}", status_code=303)


@router.post("/reset-grading")
def reset_grading(request: Request, user: dict = admin_dep):
    """Đặt lại các hồ sơ kẹt ở 'đang chấm' quá lâu (Cloud Run cắt CPU) để chấm lại."""
    n = reset_stuck_grading(request.app.state.store, user)
    return RedirectResponse(f"/admin?reset_grading={n}", status_code=303)


@router.post("/remind")
def remind(request: Request, user: dict = admin_dep):
    sent = send_reminders(request.app.state.store)
    audit.log(request.app.state.store, user, "send_reminders", "users", note=f"Gửi {sent} email nhắc hạn")
    return RedirectResponse(f"/admin?reminded={sent}", status_code=303)


@router.post("/publish")
def publish(request: Request, user: dict = admin_dep):
    count = publish_all(request.app.state.store, user)
    return RedirectResponse(f"/admin?published={count}", status_code=303)


# ---------- người dùng ----------

@router.get("/users")
def users_page(request: Request, page: int = 1, q: str = "", user: dict = admin_dep):
    store = request.app.state.store
    users = sorted(store.all("users"), key=lambda u: (u["role"], u.get("khoa", ""), u.get("ma_gv", "")))
    q = (q or "").strip()
    if q:
        ql = q.lower()
        fields = ("ho_ten", "email", "ma_gv", "khoa", "chuc_vu", "bo_mon")
        users = [u for u in users if any(ql in str(u.get(f) or "").lower() for f in fields)]
    total = len(users)
    per_page = 50
    pages = max(1, (total + per_page - 1) // per_page)
    page = max(1, min(page, pages))
    start = (page - 1) * per_page
    return render(request, "admin/users.html", user, users=users[start:start + per_page],
                  settings=get_settings(), total=total, page=page, pages=pages, per_page=per_page, q=q)


@router.post("/users/import")
async def users_import(request: Request, file: UploadFile = None, csv_text: str = Form(""), user: dict = admin_dep):
    """Import CSV: ma_gv,ho_ten,email,don_vi,bo_mon,chuc_vu,role(lecturer|council|admin),password(tùy chọn).

    Tương thích ngược: chấp nhận cột "khoa" thay cho "don_vi" và "ma_dinh_danh" thay cho "ma_gv".
    Người dùng mới không kèm password sẽ nhận mật khẩu mặc định (DEFAULT_PASSWORD, mặc định DNU@2026).
    """
    store = request.app.state.store
    settings = get_settings()
    raw = ""
    if file is not None and file.filename:
        raw = (await file.read()).decode("utf-8-sig")
    elif csv_text.strip():
        raw = csv_text
    if not raw.strip():
        raise HTTPException(400, "Chưa có dữ liệu CSV")
    reader = csv.DictReader(io.StringIO(raw.strip()))
    added, updated = 0, 0
    for row in reader:
        # chuẩn hóa khóa cột về chữ thường, bỏ khoảng trắng (hỗ trợ tiêu đề có dấu cách)
        row = {(k or "").strip().lower(): v for k, v in row.items()}
        email = (row.get("email") or "").strip().lower()
        if not email:
            continue
        role = (row.get("role") or ROLE_LECTURER).strip() or ROLE_LECTURER
        # Đơn vị: nhận cột "don_vi" (mới) hoặc "khoa" (cũ); lưu nội bộ ở khóa "khoa"
        don_vi = (row.get("don_vi") or row.get("khoa") or "").strip()
        doc = {
            "email": email, "ho_ten": (row.get("ho_ten") or "").strip(),
            "ma_gv": (row.get("ma_gv") or row.get("ma_dinh_danh") or "").strip(), "khoa": don_vi,
            "bo_mon": (row.get("bo_mon") or "").strip(), "chuc_vu": (row.get("chuc_vu") or "").strip(),
            "role": role, "active": True,
        }
        pw = (row.get("password") or "").strip()
        existing = store.find_one("users", email=email)
        if existing:
            if pw:
                doc["password_hash"] = hash_password(pw)
            store.patch("users", existing["id"], doc)
            updated += 1
        else:
            doc["password_hash"] = hash_password(pw or settings.default_password)
            store.add("users", doc)
            added += 1
    audit.log(store, user, "import_users", "users", note=f"Thêm {added}, cập nhật {updated}")
    return RedirectResponse(f"/admin/users?added={added}&updated={updated}", status_code=303)


@router.post("/users/add")
def users_add(request: Request, ho_ten: str = Form(...), email: str = Form(...), ma_gv: str = Form(""),
              khoa: str = Form(""), bo_mon: str = Form(""), chuc_vu: str = Form(""),
              role: str = Form(ROLE_LECTURER), password: str = Form(...), user: dict = admin_dep):
    store = request.app.state.store
    email = email.strip().lower()
    if role not in (ROLE_LECTURER, ROLE_COUNCIL, ROLE_ADMIN):
        raise HTTPException(400, "Vai trò không hợp lệ")
    if len(password) < 6:
        raise HTTPException(400, "Mật khẩu tối thiểu 6 ký tự")
    if store.find_one("users", email=email):
        raise HTTPException(400, "Email đã tồn tại")
    store.add("users", {
        "email": email, "ho_ten": ho_ten.strip(), "ma_gv": ma_gv.strip(),
        "khoa": khoa.strip(), "bo_mon": bo_mon.strip(), "chuc_vu": chuc_vu.strip(),
        "role": role, "active": True, "password_hash": hash_password(password),
    })
    audit.log(store, user, "add_user", f"users/{email}", note=f"role={role}")
    return RedirectResponse("/admin/users?added=1&updated=0", status_code=303)


@router.post("/users/{uid}/password")
def users_set_password(uid: str, request: Request, new_password: str = Form(...), user: dict = admin_dep):
    store = request.app.state.store
    target = store.get("users", uid)
    if not target:
        raise HTTPException(404)
    if len(new_password) < 6:
        raise HTTPException(400, "Mật khẩu tối thiểu 6 ký tự")
    set_password(store, target, new_password)
    audit.log(store, user, "reset_password", f"users/{target['email']}")
    return RedirectResponse("/admin/users?pwset=1", status_code=303)


@router.post("/users/{uid}/toggle")
def users_toggle(uid: str, request: Request, user: dict = admin_dep):
    store = request.app.state.store
    target = store.get("users", uid)
    if not target:
        raise HTTPException(404)
    new_active = not target.get("active", True)
    store.patch("users", uid, {"active": new_active})
    audit.log(store, user, "toggle_user", f"users/{target['email']}", after={"active": new_active})
    return RedirectResponse("/admin/users", status_code=303)


# ---------- cấu hình ----------

@router.get("/config")
def config_page(request: Request, user: dict = admin_dep):
    from app.services.ai_config import get_ai_config
    from app.services.ai_usage import get_stats

    store = request.app.state.store
    graded_count = sum(1 for s in store.all("submissions")
                       if s.get("ai_graded") or s.get("status") in ("graded", "approved", "published"))
    return render(request, "admin/config.html", user, timeline=get_timeline(store),
                  rubric=get_rubric(store), rubric_types=rubric_types(store),
                  seed_version=load_rubric_seed().get("version", ""),
                  graded_count=graded_count, settings=get_settings(),
                  ai_cfg=get_ai_config(store), ai_stats=get_stats(store))


@router.get("/rubric.xlsx")
def rubric_xlsx(request: Request, user: dict = admin_dep):
    from fastapi.responses import Response

    from app.services.rubric_export import rubric_to_xlsx

    data = rubric_to_xlsx(get_rubric(request.app.state.store))
    return Response(
        data, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":
                 f"attachment; filename={get_settings().org_short}-Rubric-HoiDong-{now_vn():%Y%m%d}.xlsx"},
    )


@router.get("/rubric.docx")
def rubric_docx(request: Request, user: dict = admin_dep):
    from fastapi.responses import Response

    from app.services.rubric_export import rubric_to_docx

    data = rubric_to_docx(get_rubric(request.app.state.store))
    return Response(
        data, media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition":
                 f"attachment; filename={get_settings().org_short}-Rubric-HoiDong-{now_vn():%Y%m%d}.docx"},
    )


@router.post("/rubric/reload")
def rubric_reload(request: Request, user: dict = admin_dep):
    """Nạp lại rubric mới nhất kèm theo phần mềm (ghi đè config/rubric).

    An toàn: chỉ cập nhật rubric chấm điểm, không xóa hồ sơ/điểm/người dùng.
    """
    from urllib.parse import quote

    store = request.app.state.store
    old_v, new_v = reload_rubric(store)
    audit.log(store, user, "reload_rubric", "config/rubric",
              before={"version": old_v}, after={"version": new_v},
              note=f"Nạp lại rubric {old_v or '(chưa có)'} → {new_v}")
    msg = (f"Đã nạp lại rubric mới nhất: {old_v or '(chưa có)'} → {new_v}."
           if old_v != new_v else f"Rubric đã là bản mới nhất ({new_v}).")
    return RedirectResponse(f"/admin/config?rubric_msg={quote(msg)}", status_code=303)


@router.post("/config/timeline")
def config_timeline(request: Request, deadline: str = Form(...), open_at: str = Form(...), user: dict = admin_dep):
    store = request.app.state.store
    tl = get_timeline(store)
    before = {"deadline": tl["deadline"], "open_at": tl["open_at"]}
    tl.update({"deadline": deadline.strip(), "open_at": open_at.strip()})
    # Gia hạn (hạn mới ở tương lai) → cho cron khóa lại đúng hạn mới. Hồ sơ ĐANG khóa
    # vẫn cần bấm "Mở khóa" riêng để giảng viên nộp lại.
    try:
        if parse_dt(tl["deadline"]) > now_vn():
            tl["locked_done"] = False
    except Exception:  # noqa: BLE001 — định dạng ngày sai không được chặn lưu
        pass
    store.put("config", "timeline", tl)
    audit.log(store, user, "update_timeline", "config/timeline", before=before, after=tl)
    return RedirectResponse("/admin/config?saved=1", status_code=303)


# ---------- cấu hình AI (Admin nạp API key + model) ----------

@router.post("/config/ai")
def config_ai(request: Request, api_key: str = Form(""), model: str = Form(""),
              grader: str = Form("auto"), clear_key: str = Form(""), user: dict = admin_dep):
    from app.services.ai_config import set_ai_config

    store = request.app.state.store
    set_ai_config(store, api_key=api_key, model=model, grader=grader, clear_key=bool(clear_key))
    audit.log(store, user, "update_ai_config", "config/ai",
              note=f"grader={grader}, model={model or '(giữ nguyên)'}, "
                   f"{'xóa key' if clear_key else ('cập nhật key' if api_key.strip() else 'giữ key')}")
    return RedirectResponse("/admin/config?ai=1", status_code=303)


@router.post("/config/ai/test")
def config_ai_test(request: Request, user: dict = admin_dep):
    """Gọi thử Claude để kiểm tra API key + model, tự thử lại khi máy chủ quá tải (529)."""
    from app.services.ai_config import get_ai_config
    from app.services.ai_usage import record_usage
    from app.services.grading.graders import friendly_ai_error, is_transient_error

    store = request.app.state.store
    cfg = get_ai_config(store)
    if cfg["grader"] != "claude" or not cfg["has_key"]:
        return RedirectResponse("/admin/config?ai_test=Chưa+cấu+hình+API+key+Claude", status_code=303)

    import anthropic

    client = anthropic.Anthropic(api_key=cfg["api_key"], max_retries=2)
    msg, last_exc, attempts = "", None, 4
    for attempt in range(attempts):
        try:
            resp = client.messages.create(
                model=cfg["model"], max_tokens=16,
                messages=[{"role": "user", "content": "Trả lời đúng một từ: OK"}],
            )
            record_usage(store, cfg["model"], getattr(resp, "usage", None), kind="test")
            msg = "Kết nối Claude thành công ✓"
            break
        except Exception as exc:  # noqa: BLE001 — phân loại tạm thời / cấu hình
            last_exc = exc
            if not is_transient_error(exc) or attempt == attempts - 1:
                break
            time.sleep(2 * (attempt + 1))  # backoff 2s, 4s, 6s khi Claude quá tải
    if not msg:
        msg = friendly_ai_error(last_exc, cfg["model"])
    audit.log(store, user, "test_ai", "config/ai", note=msg[:200])
    from urllib.parse import quote

    return RedirectResponse(f"/admin/config?ai_test={quote(msg)}", status_code=303)


@router.get("/ai-usage")
def ai_usage_page(request: Request, user: dict = admin_dep):
    from app.services.ai_config import get_ai_config
    from app.services.ai_usage import USD_TO_VND, get_stats

    store = request.app.state.store
    return render(request, "admin/ai_usage.html", user,
                  stats=get_stats(store), ai_cfg=get_ai_config(store), usd_vnd=USD_TO_VND)


@router.post("/ai-usage/reset")
def ai_usage_reset(request: Request, user: dict = admin_dep):
    from app.services.ai_usage import reset_stats

    store = request.app.state.store
    n = reset_stats(store)
    audit.log(store, user, "reset_ai_usage", "ai_usage", note=f"Xóa {n} bản ghi")
    return RedirectResponse("/admin/ai-usage?reset=1", status_code=303)


@router.get("/audit")
def audit_page(request: Request, user: dict = admin_dep):
    store = request.app.state.store
    logs = sorted(store.all("audit_logs"), key=lambda x: x["ts"], reverse=True)[:300]
    return render(request, "admin/audit.html", user, logs=logs)


@router.get("/emails")
def emails_page(request: Request, user: dict = admin_dep):
    store = request.app.state.store
    emails = sorted(store.all("email_logs"), key=lambda x: x["ts"], reverse=True)[:200]
    return render(request, "admin/emails.html", user, emails=emails)


# ---------- tải sản phẩm (ZIP) ----------

@router.get("/downloads")
def downloads_page(request: Request, khoa: str = "", sort: str = "ma_gv", page: int = 1,
                   user: dict = admin_dep):
    store = request.app.state.store
    users = {u["id"]: u for u in store.all("users")}
    rows = []
    for s in store.all("submissions"):
        u = users.get(s["user_id"])
        if not u or s.get("status") == "draft":
            continue
        if khoa and (u.get("khoa") or "") != khoa:
            continue
        items = store.find("submission_items", submission_id=s["id"])
        ups = [i.get("uploaded_at") for i in items if i.get("uploaded_at")]
        last_upload = max(ups) if ups else (s.get("submitted_at") or "")
        rows.append({
            "sub": s, "user": u,
            "n_files": sum(1 for i in items if i.get("type") == "file"),
            "n_links": sum(1 for i in items if i.get("type") == "link"),
            "last_upload": last_upload,
        })
    sorters = {
        "ma_gv": (lambda r: r["user"].get("ma_gv", ""), False),
        "don_vi": (lambda r: (r["user"].get("khoa", ""), r["user"].get("ma_gv", "")), False),
        "ho_ten": (lambda r: r["user"].get("ho_ten", ""), False),
        "upload": (lambda r: r["last_upload"] or "", True),   # mới nhất trước
        "files": (lambda r: r["n_files"], True),
    }
    key, reverse = sorters.get(sort, sorters["ma_gv"])
    rows.sort(key=key, reverse=reverse)
    total = len(rows)
    per_page = 50
    pages = max(1, (total + per_page - 1) // per_page)
    page = max(1, min(page, pages))
    start = (page - 1) * per_page
    khoas = sorted({u.get("khoa", "") for u in users.values() if u.get("khoa")})
    return render(request, "admin/downloads.html", user, rows=rows[start:start + per_page], khoas=khoas,
                  khoa=khoa, sort=sort, total=total, page=page, pages=pages, per_page=per_page)


def _zip_response(gen, fname: str):
    from urllib.parse import quote

    from fastapi.responses import StreamingResponse

    return StreamingResponse(gen, media_type="application/zip", headers={
        "Content-Disposition": f"attachment; filename*=UTF-8''{quote(fname)}",
        "X-Accel-Buffering": "no",  # không đệm, đẩy luồng ngay
    })


@router.get("/download/all.zip")
def download_all(request: Request, khoa: str = "", user: dict = admin_dep):
    from app.services.downloads import stream_zip, zip_filename

    store, storage = request.app.state.store, request.app.state.storage
    subs = [s for s in store.all("submissions")
            if s.get("status") != "draft" and store.get("users", s["user_id"])
            and (not khoa or (store.get("users", s["user_id"]).get("khoa") or "") == khoa)]
    audit.log(store, user, "download_all", "submissions", note=f"khoa={khoa or 'tất cả'} ({len(subs)} hồ sơ)")
    tag = f"-{khoa}" if khoa else ""
    return _zip_response(stream_zip(store, storage, subs, manifest=True, base_by_khoa=True), zip_filename(tag))


@router.post("/download/selected.zip")
def download_selected(request: Request, sid: list[str] = Form(default=[]), user: dict = admin_dep):
    from app.services.downloads import stream_zip, zip_filename

    store, storage = request.app.state.store, request.app.state.storage
    subs = [s for s in (store.get("submissions", i) for i in sid) if s and s.get("status") != "draft"]
    if not subs:
        raise HTTPException(400, "Chưa chọn hồ sơ nào để tải")
    audit.log(store, user, "download_selected", "submissions", note=f"{len(subs)} hồ sơ")
    return _zip_response(stream_zip(store, storage, subs, manifest=True, base_by_khoa=True),
                         zip_filename(f"-chon-{len(subs)}gv"))


# ---------- sao lưu ----------

@router.get("/backup")
def backup(request: Request, user: dict = admin_dep):
    import json
    from fastapi.responses import Response

    store = request.app.state.store
    collections = ["users", "submissions", "submission_items", "grading_runs", "scores",
                   "reviews", "appeals", "audit_logs", "config", "email_logs"]
    data = {c: store.all(c) for c in collections}
    audit.log(store, user, "backup", "all", note="Xuất bản sao lưu JSON")
    return Response(
        json.dumps(data, ensure_ascii=False, indent=1),
        media_type="application/json",
        headers={"Content-Disposition": f"attachment; filename=dnu-ai-assess-backup-{now_vn():%Y%m%d-%H%M}.json"},
    )
